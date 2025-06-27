"""
Patient Import Utility

This module provides functionality to import patient data from various formats
including CSV, Excel, and JSON.
"""

import csv
import json
from datetime import datetime, date
from typing import List, Optional, Dict, Any, Tuple
from pathlib import Path

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from src.models.complete_models import Patient
from src.core.database import db_manager


class ImportError(Exception):
    """Custom exception for import errors"""
    pass


class PatientImporter:
    """Utility class for importing patient data"""
    
    def __init__(self):
        self.supported_formats = ['csv', 'json']
        if EXCEL_AVAILABLE:
            self.supported_formats.append('xlsx')
        
        # Field mapping for CSV/Excel imports
        self.field_mapping = {
            'id': 'id',
            'first_name': 'first_name',
            'last_name': 'last_name',
            'date_of_birth': 'date_of_birth',
            # Note: 'age' is excluded as it's a computed property from date_of_birth
            'gender': 'gender',
            'phone': 'phone',
            'email': 'email',
            'address': 'address',
            'city': 'city',
            'state': 'state',
            'zip_code': 'zip_code',
            'country': 'country',
            'ssn': 'ssn',
            'id_number': 'id_number',
            'blood_type': 'blood_type',
            'height': 'height',
            'weight': 'weight',
            'insurance_provider': 'insurance_provider',
            'insurance_number': 'insurance_number',
            'insurance_group': 'insurance_group',
            'emergency_contact_name': 'emergency_contact_name',
            'emergency_contact_relationship': 'emergency_contact_relationship',
            'emergency_contact_phone': 'emergency_contact_phone',
            'allergies': 'allergies',
            'chronic_conditions': 'chronic_conditions',
            'current_medications': 'current_medications',
            'family_history': 'family_history',
            'preferred_language': 'preferred_language',
            'notes': 'notes'
        }
        
        # Alternative field names that might be used in imports
        self.field_aliases = {
            'firstname': 'first_name',
            'lastname': 'last_name',
            'dob': 'date_of_birth',
            'birth_date': 'date_of_birth',
            'telephone': 'phone',
            'phone_number': 'phone',
            'email_address': 'email',
            'street': 'address',
            'street_address': 'address',
            'postal_code': 'zip_code',
            'zipcode': 'zip_code',
            'social_security': 'ssn',
            'emergency_name': 'emergency_contact_name',
            'emergency_relationship': 'emergency_contact_relationship',
            'emergency_phone': 'emergency_contact_phone',
            'insurance': 'insurance_provider',
            'insurance_company': 'insurance_provider'
        }
    
    def normalize_headers(self, headers: List[str]) -> List[str]:
        """Normalize column headers to match field names"""
        normalized = []
        for header in headers:
            # Convert to lowercase and replace spaces/special chars with underscores
            normalized_header = header.lower().strip().replace(' ', '_').replace('-', '_')
            
            # Check if it's a direct match
            if normalized_header in self.field_mapping:
                normalized.append(normalized_header)
            # Check aliases
            elif normalized_header in self.field_aliases:
                normalized.append(self.field_aliases[normalized_header])
            else:
                # Keep original if no mapping found
                normalized.append(normalized_header)
        
        return normalized
    
    def parse_date(self, date_str: str) -> Optional[date]:
        """Parse date string into date object"""
        if not date_str or date_str.strip() == '':
            return None
        
        date_formats = [
            '%Y-%m-%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
            '%Y/%m/%d',
            '%m-%d-%Y',
            '%d-%m-%Y'
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str.strip(), fmt).date()
            except ValueError:
                continue
        
        return None
    
    def validate_patient_data(self, patient_data: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate patient data and return validation result and errors"""
        errors = []
        
        # Required fields
        if not patient_data.get('first_name'):
            errors.append("First name is required")
        
        if not patient_data.get('last_name'):
            errors.append("Last name is required")
        
        # Email validation
        email = patient_data.get('email')
        if email and '@' not in email:
            errors.append("Invalid email format")
        
        # Age validation
        age = patient_data.get('age')
        if age is not None:
            try:
                age_int = int(age)
                if age_int < 0 or age_int > 150:
                    errors.append("Age must be between 0 and 150")
            except (ValueError, TypeError):
                errors.append("Age must be a valid number")
        
        # Gender validation
        gender = patient_data.get('gender')
        if gender and gender not in ['Male', 'Female', 'Other']:
            # Try to normalize common variations
            gender_lower = gender.lower()
            if gender_lower in ['m', 'male']:
                patient_data['gender'] = 'Male'
            elif gender_lower in ['f', 'female']:
                patient_data['gender'] = 'Female'
            else:
                patient_data['gender'] = 'Other'
        
        # Height/Weight validation
        height = patient_data.get('height')
        if height is not None:
            try:
                height_int = int(height)
                if height_int < 50 or height_int > 300:
                    errors.append("Height must be between 50 and 300 cm")
            except (ValueError, TypeError):
                errors.append("Height must be a valid number")
        
        weight = patient_data.get('weight')
        if weight is not None:
            try:
                weight_int = int(weight)
                if weight_int < 1 or weight_int > 500:
                    errors.append("Weight must be between 1 and 500 kg")
            except (ValueError, TypeError):
                errors.append("Weight must be a valid number")
        
        return len(errors) == 0, errors
    
    def import_from_csv(self, filename: str) -> Tuple[int, int, List[str]]:
        """Import patients from CSV file"""
        imported_count = 0
        error_count = 0
        errors = []
        
        try:
            with open(filename, 'r', encoding='utf-8') as csvfile:
                # Detect delimiter
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter
                
                reader = csv.DictReader(csvfile, delimiter=delimiter)
                
                # Normalize headers
                fieldnames = self.normalize_headers(reader.fieldnames)
                reader.fieldnames = fieldnames
                
                with db_manager.get_session() as session:
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            # Clean and prepare data
                            patient_data = {}
                            
                            for field, value in row.items():
                                if field in self.field_mapping and value and str(value).strip():
                                    clean_value = str(value).strip()
                                    
                                    # Special handling for specific fields
                                    if field == 'date_of_birth':
                                        patient_data[field] = self.parse_date(clean_value)
                                    elif field in ['age', 'height', 'weight']:
                                        try:
                                            patient_data[field] = int(clean_value) if clean_value else None
                                        except ValueError:
                                            patient_data[field] = None
                                    else:
                                        patient_data[field] = clean_value
                            
                            # Validate data
                            is_valid, validation_errors = self.validate_patient_data(patient_data)
                            
                            if not is_valid:
                                error_count += 1
                                errors.append(f"Row {row_num}: {'; '.join(validation_errors)}")
                                continue
                            
                            # Check if patient already exists (by name and DOB or email)
                            existing_patient = None
                            if patient_data.get('email'):
                                existing_patient = session.query(Patient).filter(
                                    Patient.email == patient_data['email']
                                ).first()
                            
                            if not existing_patient and patient_data.get('first_name') and patient_data.get('last_name'):
                                existing_patient = session.query(Patient).filter(
                                    Patient.first_name == patient_data['first_name'],
                                    Patient.last_name == patient_data['last_name'],
                                    Patient.date_of_birth == patient_data.get('date_of_birth')
                                ).first()
                            
                            if existing_patient:
                                error_count += 1
                                errors.append(f"Row {row_num}: Patient already exists")
                                continue
                            
                            # Create new patient
                            patient = Patient()
                            
                            # Set fields (excluding computed properties like 'age')
                            for field, value in patient_data.items():
                                if field != 'age' and hasattr(patient, field) and value is not None:
                                    setattr(patient, field, value)
                            
                            # Calculate age if not provided but DOB is available
                            # Note: Age is automatically calculated as a property from date_of_birth
                            # No need to set it manually
                            
                            # Set creation date
                            patient.created_at = datetime.now()
                            
                            session.add(patient)
                            imported_count += 1
                            
                        except Exception as e:
                            error_count += 1
                            errors.append(f"Row {row_num}: {str(e)}")
                    
                    # Commit all changes
                    session.commit()
            
        except Exception as e:
            errors.append(f"File reading error: {str(e)}")
            return 0, 0, errors
        
        return imported_count, error_count, errors
    
    def import_from_excel(self, filename: str) -> Tuple[int, int, List[str]]:
        """Import patients from Excel file"""
        if not EXCEL_AVAILABLE:
            return 0, 0, ["Excel import not available. Please install openpyxl."]
        
        imported_count = 0
        error_count = 0
        errors = []
        
        try:
            workbook = openpyxl.load_workbook(filename)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value))
                else:
                    break
            
            # Normalize headers
            normalized_headers = self.normalize_headers(headers)
            
            with db_manager.get_session() as session:
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Create patient data dictionary
                        patient_data = {}
                        
                        for i, value in enumerate(row[:len(normalized_headers)]):
                            field = normalized_headers[i]
                            
                            if field in self.field_mapping and value is not None:
                                clean_value = str(value).strip() if value else None
                                
                                if clean_value:
                                    # Special handling for specific fields
                                    if field == 'date_of_birth':
                                        if isinstance(value, date):
                                            patient_data[field] = value
                                        else:
                                            patient_data[field] = self.parse_date(clean_value)
                                    elif field in ['age', 'height', 'weight']:
                                        try:
                                            patient_data[field] = int(float(clean_value)) if clean_value else None
                                        except ValueError:
                                            patient_data[field] = None
                                    else:
                                        patient_data[field] = clean_value
                        
                        # Skip empty rows
                        if not any(patient_data.values()):
                            continue
                        
                        # Validate data
                        is_valid, validation_errors = self.validate_patient_data(patient_data)
                        
                        if not is_valid:
                            error_count += 1
                            errors.append(f"Row {row_num}: {'; '.join(validation_errors)}")
                            continue
                        
                        # Check if patient already exists
                        existing_patient = None
                        if patient_data.get('email'):
                            existing_patient = session.query(Patient).filter(
                                Patient.email == patient_data['email']
                            ).first()
                        
                        if not existing_patient and patient_data.get('first_name') and patient_data.get('last_name'):
                            existing_patient = session.query(Patient).filter(
                                Patient.first_name == patient_data['first_name'],
                                Patient.last_name == patient_data['last_name'],
                                Patient.date_of_birth == patient_data.get('date_of_birth')
                            ).first()
                        
                        if existing_patient:
                            error_count += 1
                            errors.append(f"Row {row_num}: Patient already exists")
                            continue
                        
                        # Create new patient
                        patient = Patient()
                        
                        # Set fields (excluding computed properties like 'age')
                        for field, value in patient_data.items():
                            if field != 'age' and hasattr(patient, field) and value is not None:
                                setattr(patient, field, value)
                        
                        # Age is automatically calculated as a property from date_of_birth
                        # No need to set it manually
                        
                        patient.created_at = datetime.now()
                        
                        session.add(patient)
                        imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")
                
                session.commit()
            
        except Exception as e:
            errors.append(f"File reading error: {str(e)}")
            return 0, 0, errors
        
        return imported_count, error_count, errors
    
    def import_from_json(self, filename: str) -> Tuple[int, int, List[str]]:
        """Import patients from JSON file"""
        imported_count = 0
        error_count = 0
        errors = []
        
        try:
            with open(filename, 'r', encoding='utf-8') as jsonfile:
                data = json.load(jsonfile)
            
            # Handle different JSON formats
            patients_data = []
            if isinstance(data, list):
                patients_data = data
            elif isinstance(data, dict):
                if 'patients' in data:
                    patients_data = data['patients']
                else:
                    patients_data = [data]
            
            with db_manager.get_session() as session:
                for i, patient_data in enumerate(patients_data):
                    try:
                        # Parse date fields
                        if 'date_of_birth' in patient_data and patient_data['date_of_birth']:
                            if isinstance(patient_data['date_of_birth'], str):
                                patient_data['date_of_birth'] = datetime.fromisoformat(
                                    patient_data['date_of_birth'].replace('Z', '+00:00')
                                ).date()
                        
                        # Validate data
                        is_valid, validation_errors = self.validate_patient_data(patient_data)
                        
                        if not is_valid:
                            error_count += 1
                            errors.append(f"Patient {i + 1}: {'; '.join(validation_errors)}")
                            continue
                        
                        # Check if patient already exists
                        existing_patient = None
                        if patient_data.get('email'):
                            existing_patient = session.query(Patient).filter(
                                Patient.email == patient_data['email']
                            ).first()
                        
                        if not existing_patient and patient_data.get('first_name') and patient_data.get('last_name'):
                            existing_patient = session.query(Patient).filter(
                                Patient.first_name == patient_data['first_name'],
                                Patient.last_name == patient_data['last_name'],
                                Patient.date_of_birth == patient_data.get('date_of_birth')
                            ).first()
                        
                        if existing_patient:
                            error_count += 1
                            errors.append(f"Patient {i + 1}: Patient already exists")
                            continue
                        
                        # Create new patient
                        patient = Patient()
                        
                        # Set fields (excluding computed properties like 'age')
                        for field, value in patient_data.items():
                            if field != 'age' and hasattr(patient, field) and value is not None:
                                setattr(patient, field, value)
                        
                        # Age is automatically calculated as a property from date_of_birth
                        # No need to set it manually
                        
                        if not patient.created_at:
                            patient.created_at = datetime.now()
                        
                        session.add(patient)
                        imported_count += 1
                        
                    except Exception as e:
                        error_count += 1
                        errors.append(f"Patient {i + 1}: {str(e)}")
                
                session.commit()
            
        except Exception as e:
            errors.append(f"File reading error: {str(e)}")
            return 0, 0, errors
        
        return imported_count, error_count, errors
    
    def import_patients(self, filename: str) -> Tuple[int, int, List[str]]:
        """Import patients from file (auto-detect format)"""
        file_path = Path(filename)
        file_extension = file_path.suffix.lower()
        
        if file_extension == '.csv':
            return self.import_from_csv(filename)
        elif file_extension in ['.xlsx', '.xls']:
            return self.import_from_excel(filename)
        elif file_extension == '.json':
            return self.import_from_json(filename)
        else:
            return 0, 0, [f"Unsupported file format: {file_extension}"]
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported import formats"""
        return self.supported_formats.copy()
    
    def preview_import(self, filename: str, max_rows: int = 10) -> Tuple[List[Dict], List[str]]:
        """Preview import data without actually importing"""
        file_path = Path(filename)
        file_extension = file_path.suffix.lower()
        
        preview_data = []
        errors = []
        
        try:
            if file_extension == '.csv':
                with open(filename, 'r', encoding='utf-8') as csvfile:
                    sample = csvfile.read(1024)
                    csvfile.seek(0)
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample).delimiter
                    
                    reader = csv.DictReader(csvfile, delimiter=delimiter)
                    fieldnames = self.normalize_headers(reader.fieldnames)
                    reader.fieldnames = fieldnames
                    
                    for i, row in enumerate(reader):
                        if i >= max_rows:
                            break
                        preview_data.append(dict(row))
            
            elif file_extension in ['.xlsx', '.xls'] and EXCEL_AVAILABLE:
                workbook = openpyxl.load_workbook(filename)
                worksheet = workbook.active
                
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(str(cell.value))
                    else:
                        break
                
                normalized_headers = self.normalize_headers(headers)
                
                for i, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True)):
                    if i >= max_rows:
                        break
                    
                    row_data = {}
                    for j, value in enumerate(row[:len(normalized_headers)]):
                        if j < len(normalized_headers):
                            row_data[normalized_headers[j]] = str(value) if value is not None else ''
                    
                    preview_data.append(row_data)
            
            elif file_extension == '.json':
                with open(filename, 'r', encoding='utf-8') as jsonfile:
                    data = json.load(jsonfile)
                
                patients_data = []
                if isinstance(data, list):
                    patients_data = data[:max_rows]
                elif isinstance(data, dict):
                    if 'patients' in data:
                        patients_data = data['patients'][:max_rows]
                    else:
                        patients_data = [data]
                
                preview_data = patients_data
            
            else:
                errors.append(f"Unsupported file format: {file_extension}")
        
        except Exception as e:
            errors.append(f"Error reading file: {str(e)}")
        
        return preview_data, errors
