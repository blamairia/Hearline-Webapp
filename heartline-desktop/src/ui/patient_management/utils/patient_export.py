"""
Patient Export Utility

This module provides functionality to export patient data in various formats
including CSV, Excel, PDF, and JSON.
"""

import csv
import json
from datetime import datetime
from typing import List, Optional, Dict, Any
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from src.models.complete_models import Patient


class PatientExporter:
    """Utility class for exporting patient data"""
    
    def __init__(self):
        self.supported_formats = ['csv', 'json']
        if EXCEL_AVAILABLE:
            self.supported_formats.append('xlsx')
        if PDF_AVAILABLE:
            self.supported_formats.append('pdf')
    
    def export_to_csv(self, patients: List[Patient], filename: str) -> bool:
        """Export patients to CSV format"""
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers
                headers = [
                    'ID', 'First Name', 'Last Name', 'Date of Birth', 'Age', 'Gender',
                    'Phone', 'Email', 'Address', 'City', 'State', 'ZIP Code', 'Country',
                    'SSN', 'ID Number', 'Blood Type', 'Height (cm)', 'Weight (kg)',
                    'Insurance Provider', 'Insurance Number', 'Insurance Group',
                    'Emergency Contact Name', 'Emergency Contact Relationship', 'Emergency Contact Phone',
                    'Allergies', 'Chronic Conditions', 'Current Medications', 'Family History',
                    'Preferred Language', 'Notes', 'Created Date', 'Status'
                ]
                writer.writerow(headers)
                
                # Write patient data
                for patient in patients:
                    row = [
                        patient.id,
                        patient.first_name or '',
                        patient.last_name or '',
                        patient.date_of_birth.strftime('%Y-%m-%d') if patient.date_of_birth else '',
                        patient.age or '',
                        patient.gender or '',
                        patient.phone or '',
                        patient.email or '',
                        patient.address or '',
                        patient.city or '',
                        patient.state or '',
                        patient.zip_code or '',
                        patient.country or '',
                        patient.ssn or '',
                        patient.id_number or '',
                        patient.blood_type or '',
                        patient.height or '',
                        patient.weight or '',
                        patient.insurance_provider or '',
                        patient.insurance_number or '',
                        patient.insurance_group or '',
                        patient.emergency_contact_name or '',
                        patient.emergency_contact_relationship or '',
                        patient.emergency_contact_phone or '',
                        patient.allergies or '',
                        patient.chronic_conditions or '',
                        patient.current_medications or '',
                        patient.family_history or '',
                        patient.preferred_language or '',
                        patient.notes or '',
                        patient.created_at.strftime('%Y-%m-%d %H:%M:%S') if patient.created_at else '',
                        'Active' if getattr(patient, 'is_active', True) else 'Inactive'
                    ]
                    writer.writerow(row)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to CSV: {str(e)}")
            return False
    
    def export_to_json(self, patients: List[Patient], filename: str) -> bool:
        """Export patients to JSON format"""
        try:
            patients_data = []
            
            for patient in patients:
                patient_dict = {
                    'id': patient.id,
                    'first_name': patient.first_name,
                    'last_name': patient.last_name,
                    'date_of_birth': patient.date_of_birth.isoformat() if patient.date_of_birth else None,
                    'age': patient.age,
                    'gender': patient.gender,
                    'phone': patient.phone,
                    'email': patient.email,
                    'address': patient.address,
                    'city': patient.city,
                    'state': patient.state,
                    'zip_code': patient.zip_code,
                    'country': patient.country,
                    'ssn': patient.ssn,
                    'id_number': patient.id_number,
                    'blood_type': patient.blood_type,
                    'height': patient.height,
                    'weight': patient.weight,
                    'insurance_provider': patient.insurance_provider,
                    'insurance_number': patient.insurance_number,
                    'insurance_group': patient.insurance_group,
                    'emergency_contact_name': patient.emergency_contact_name,
                    'emergency_contact_relationship': patient.emergency_contact_relationship,
                    'emergency_contact_phone': patient.emergency_contact_phone,
                    'allergies': patient.allergies,
                    'chronic_conditions': patient.chronic_conditions,
                    'current_medications': patient.current_medications,
                    'family_history': patient.family_history,
                    'preferred_language': patient.preferred_language,
                    'notes': patient.notes,
                    'created_at': patient.created_at.isoformat() if patient.created_at else None,
                    'is_active': getattr(patient, 'is_active', True)
                }
                patients_data.append(patient_dict)
            
            export_data = {
                'export_date': datetime.now().isoformat(),
                'total_patients': len(patients),
                'patients': patients_data
            }
            
            with open(filename, 'w', encoding='utf-8') as jsonfile:
                json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            print(f"Error exporting to JSON: {str(e)}")
            return False
    
    def export_to_excel(self, patients: List[Patient], filename: str) -> bool:
        """Export patients to Excel format"""
        if not EXCEL_AVAILABLE:
            print("Excel export not available. Please install openpyxl.")
            return False
        
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Patients"
            
            # Define headers
            headers = [
                'ID', 'First Name', 'Last Name', 'Date of Birth', 'Age', 'Gender',
                'Phone', 'Email', 'Address', 'City', 'State', 'ZIP Code', 'Country',
                'Insurance Provider', 'Emergency Contact', 'Allergies', 'Created Date'
            ]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Write patient data
            for row, patient in enumerate(patients, 2):
                data = [
                    patient.id,
                    patient.first_name or '',
                    patient.last_name or '',
                    patient.date_of_birth.strftime('%Y-%m-%d') if patient.date_of_birth else '',
                    patient.age or '',
                    patient.gender or '',
                    patient.phone or '',
                    patient.email or '',
                    patient.address or '',
                    patient.city or '',
                    patient.state or '',
                    patient.zip_code or '',
                    patient.country or '',
                    patient.insurance_provider or '',
                    patient.emergency_contact_name or '',
                    patient.allergies or '',
                    patient.created_at.strftime('%Y-%m-%d') if patient.created_at else ''
                ]
                
                for col, value in enumerate(data, 1):
                    worksheet.cell(row=row, column=col, value=value)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(filename)
            return True
            
        except Exception as e:
            print(f"Error exporting to Excel: {str(e)}")
            return False
    
    def export_to_pdf(self, patients: List[Patient], filename: str) -> bool:
        """Export patients to PDF format"""
        if not PDF_AVAILABLE:
            print("PDF export not available. Please install reportlab.")
            return False
        
        try:
            doc = SimpleDocTemplate(filename, pagesize=A4)
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            
            # Title
            title = Paragraph("Patient List Report", title_style)
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Summary
            summary_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
            summary_text += f"Total Patients: {len(patients)}<br/><br/>"
            summary = Paragraph(summary_text, styles['Normal'])
            elements.append(summary)
            
            # Table data
            table_data = [
                ['ID', 'Name', 'Age', 'Gender', 'Phone', 'Insurance']
            ]
            
            for patient in patients:
                full_name = f"{patient.first_name or ''} {patient.last_name or ''}".strip()
                row = [
                    str(patient.id),
                    full_name,
                    str(patient.age) if patient.age else 'N/A',
                    patient.gender or 'N/A',
                    patient.phone or 'N/A',
                    patient.insurance_provider or 'None'
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            return True
            
        except Exception as e:
            print(f"Error exporting to PDF: {str(e)}")
            return False
    
    def export_patients(self, patients: List[Patient], filename: str, format_type: str) -> bool:
        """Export patients in the specified format"""
        format_type = format_type.lower()
        
        if format_type not in self.supported_formats:
            print(f"Unsupported format: {format_type}")
            return False
        
        if format_type == 'csv':
            return self.export_to_csv(patients, filename)
        elif format_type == 'json':
            return self.export_to_json(patients, filename)
        elif format_type == 'xlsx':
            return self.export_to_excel(patients, filename)
        elif format_type == 'pdf':
            return self.export_to_pdf(patients, filename)
        else:
            return False
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        return self.supported_formats.copy()
    
    def create_summary_report(self, patients: List[Patient]) -> Dict[str, Any]:
        """Create a summary report of patient data"""
        if not patients:
            return {
                'total_patients': 0,
                'demographics': {},
                'insurance_status': {},
                'medical_info': {}
            }
        
        # Demographics
        gender_counts = {}
        age_ranges = {'0-18': 0, '19-35': 0, '36-50': 0, '51-65': 0, '65+': 0}
        
        # Insurance
        insured_count = 0
        uninsured_count = 0
        
        # Medical
        patients_with_allergies = 0
        patients_with_chronic_conditions = 0
        
        for patient in patients:
            # Gender
            gender = patient.gender or 'Unknown'
            gender_counts[gender] = gender_counts.get(gender, 0) + 1
            
            # Age
            if patient.age:
                if patient.age <= 18:
                    age_ranges['0-18'] += 1
                elif patient.age <= 35:
                    age_ranges['19-35'] += 1
                elif patient.age <= 50:
                    age_ranges['36-50'] += 1
                elif patient.age <= 65:
                    age_ranges['51-65'] += 1
                else:
                    age_ranges['65+'] += 1
            
            # Insurance
            if patient.insurance_provider:
                insured_count += 1
            else:
                uninsured_count += 1
            
            # Medical conditions
            if patient.allergies:
                patients_with_allergies += 1
            if patient.chronic_conditions:
                patients_with_chronic_conditions += 1
        
        return {
            'total_patients': len(patients),
            'demographics': {
                'gender_distribution': gender_counts,
                'age_distribution': age_ranges
            },
            'insurance_status': {
                'insured': insured_count,
                'uninsured': uninsured_count,
                'insurance_rate': round((insured_count / len(patients)) * 100, 1) if patients else 0
            },
            'medical_info': {
                'patients_with_allergies': patients_with_allergies,
                'patients_with_chronic_conditions': patients_with_chronic_conditions,
                'allergy_rate': round((patients_with_allergies / len(patients)) * 100, 1) if patients else 0,
                'chronic_condition_rate': round((patients_with_chronic_conditions / len(patients)) * 100, 1) if patients else 0
            }
        }
